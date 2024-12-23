import os.path
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv
from conftest import create_zip_file


def test_create_zip_file(create_zip_file):
    assert os.path.isfile('zip_file.zip')


def test_pdf(create_zip_file):
    with ZipFile('zip_file.zip') as zip_file:
        with zip_file.open('AIThatWritesEssays.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            assert "highly skilled" in page.extract_text()


def test_xlsx(create_zip_file):
    with ZipFile('zip_file.zip') as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert sheet.cell(row=2, column=3).value == 'Abril'


def test_csv(create_zip_file):
    with ZipFile('zip_file.zip') as zip_file:
        with zip_file.open('file_example_csv.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            first_row = csvreader[0]
            assert first_row[1] == 'First Name'
